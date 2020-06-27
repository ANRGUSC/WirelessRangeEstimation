from estimate_distance_matrix import estimate_distance_matrix
from make_simulated_data import SimulateRssTrial, SimulateTrialsFromTrinityData

import re
import time
import json
import pickle
import os
import sys
import numpy as np
import pandas as pd
import multiprocessing
from openpyxl import load_workbook
import matplotlib.pyplot as plt
from scipy.special import ndtr

def GetNodeNumFromFilepath(fpath):
    temp = fpath[fpath.index("_data/")+6:fpath.index("nodes_")]
    temp = re.compile(r'collection/.*').search(fpath)
    if temp is None:
        temp = re.compile(r'_data/.*').search(fpath).group(0)
        temp = re.compile(r'/.*nodes').search(temp).group(0)
        temp = temp[1:-5]
    else:
        temp = re.compile(r'/.*nodes').search(temp.group(0)).group(0)
        temp = temp[1:-5]

    return int(temp)

def GetConfusionInfo(true_dist_upper, est_dist_upper, threshold):
    pos_locs = np.argwhere(true_dist_upper < threshold)
    neg_locs = np.argwhere(true_dist_upper > threshold)
    true_pos = 0
    false_neg = 0
    for loc in pos_locs:
        if est_dist_upper[loc] < threshold:
            true_pos += 1
        else:
            false_neg += 1
    true_neg = 0
    false_pos = 0
    for loc in neg_locs:
        if est_dist_upper[loc] > threshold:
            true_neg += 1
        else:
            false_pos += 1

    if true_pos+false_neg > 0:
        true_pos_rate = float(true_pos)/float(true_pos+false_neg)
    else:
        true_pos_rate = 1
    if true_neg+false_pos > 0:
        false_pos_rate = float(false_pos)/float(true_neg+false_pos)
    else:
        false_pos_rate = 1

    return true_pos_rate, false_pos_rate

def CharacterizePerformance(filepath, approaches, threshold):
    true_dist_df = pd.read_excel(filepath, sheet_name="true_dist", index_col=0)
    runtime_df = pd.read_excel(filepath, sheet_name="runtimes", index_col=0)
    true_dist_arr = true_dist_df.to_numpy()
    assert(true_dist_arr.shape[0] == true_dist_arr.shape[1])
    n = true_dist_arr.shape[0]

    # extracting upper diagonal part of matrices because of symmetry
    true_dist_upper = true_dist_arr[np.triu_indices(n, k=1)]

    num_nodes = GetNodeNumFromFilepath(filepath)
    with open('sim_data_params.json') as f:
        sim_data_params = json.load(f)
    spring_model_params = sim_data_params["spring_params"]
    n_init = spring_model_params[4]
    large_data = num_nodes >= sim_data_params["large_data_thresh"]

    results = []
    for approach in approaches:

        if large_data and "sdp" in approach:
            continue
        if approach == "spring_model":
            app_name = "spring_model_%dinits"%(n_init)
        else:
            app_name = approach

        est_dist_df = pd.read_excel(filepath, sheet_name = app_name+"_dist", index_col=0)
        est_dist_arr = est_dist_df.to_numpy()
        assert(true_dist_arr.shape == est_dist_arr.shape)

        # extracting upper diagonal part of matrices because of symmetry
        est_dist_upper = est_dist_arr[np.triu_indices(n, k=1)]
        abs_diff_upper = np.abs(true_dist_upper - est_dist_upper)
        percent_error_upper = abs_diff_upper/true_dist_upper

        avg_error = abs_diff_upper.mean()
        avg_percent_error = percent_error_upper.mean()
        max_error = abs_diff_upper.max()
        max_percent_error = percent_error_upper.max()
        true_pos_rate, false_pos_rate = GetConfusionInfo(true_dist_upper, est_dist_upper, threshold)
        timing = runtime_df[app_name][0]
        results.append([app_name, max_error, max_percent_error, avg_error, avg_percent_error, true_pos_rate, false_pos_rate, timing])

    perf_df = pd.DataFrame(results, columns=["estimation_technique", "max_error", "max_percent_error", "avg_error", "avg_percent_error", "true_pos_rate", "false_pos_rate", "runtime"])

    book = load_workbook(filepath)
    for sheet in book.sheetnames:
        if "estimation_performance" in sheet:
            tag = book.get_sheet_by_name(sheet)
            book.remove_sheet(tag)

    writer = pd.ExcelWriter(filepath, engine = 'openpyxl')
    writer.book = book
    perf_df.to_excel(writer, sheet_name = "estimation_performance")
    writer.save()
    writer.close()
    print("Performance calculated for:", filepath)

def CollectSettingData(exp_setting, experiment_paths, collection_dir):
    writer = pd.ExcelWriter(collection_dir+exp_setting+"_collection.xlsx", engine = 'openpyxl')
    for trial_path in experiment_paths:
        trial_name = trial_path[trial_path.find("_trial")+1:trial_path.find(".xlsx")]
        trial_results_df = pd.read_excel(trial_path, sheet_name="estimation_performance", index_col=0)
        trial_results_df.to_excel(writer, sheet_name = trial_name)
    writer.save()
    writer.close()
    print("Gathered results for:", exp_setting)

def GatherAllTrials(data_dir, collection_dir):
    if not (os.path.isdir(collection_dir)):
        os.mkdir(collection_dir)

    files = [data_dir+item for item in os.listdir(data_dir) if (".xlsx" in item and "trial" in item)]
    files.sort()
    setting_paths = {}
    for filepath in files:
        setting = filepath[filepath.find("_data/")+6:filepath.find("_trial")]
        if setting not in setting_paths:
            setting_paths[setting] = []
        setting_paths[setting].append(filepath)

    nproc = multiprocessing.cpu_count()
    pool = multiprocessing.Pool(nproc-2)
    for setting in setting_paths.keys():
        CollectSettingData(setting, setting_paths[setting], collection_dir)
        # pool.apply_async(CollectSettingData, args = (setting, setting_paths[setting], collection_dir))
    # pool.close()
    # pool.join()

def TestSNLApproaches(filepath, approaches, ble_params):
    start = time.time()
    num_nodes = GetNodeNumFromFilepath(filepath)
    with open('sim_data_params.json') as f:
        sim_data_params = json.load(f)
    spring_model_params = sim_data_params["spring_params"]
    n_init = spring_model_params[4]
    large_data = num_nodes >= sim_data_params["large_data_thresh"]

    # Read matrix from .xlsx and convert to numpy array
    rss_mat = pd.read_excel(filepath, sheet_name="rss_data", index_col=0).to_numpy()

    # Open current workbook so to not erase all existing data
    book = load_workbook(filepath)

    writer = pd.ExcelWriter(filepath, engine = 'openpyxl')
    writer.book = book

    true_dist_df = pd.read_excel(filepath, sheet_name="true_dist", index_col=0)
    true_dist_arr = true_dist_df.to_numpy()
    assert(true_dist_arr.shape[0] == true_dist_arr.shape[1])
    n = true_dist_arr.shape[0]

    # Run all trials and save results to .xslx file
    runtimes = {}
    for approach in approaches:
        if large_data and "sdp" in approach:
            continue

        if approach == "spring_model":
            app_name = "spring_model_%dinits"%(n_init)
        else:
            app_name = approach

        est_dist_arr, est_locs, timing = estimate_distance_matrix(rss_mat, use_model=approach, estimate_distance_params=ble_params, spring_model_params=spring_model_params)
        assert(true_dist_arr.shape == est_dist_arr.shape)

        # Clean out previous recordings
        for sheet in writer.book.sheetnames:
            if app_name+"_dist" in sheet or app_name+"_locs" in sheet or "runtimes" in sheet:
                tag = writer.book.get_sheet_by_name(sheet)
                book.remove_sheet(tag)

        runtimes[app_name] = timing
        est_dist_df = pd.DataFrame(est_dist_arr)
        est_dist_df.to_excel(writer, sheet_name = app_name+"_dist")
        if est_locs is not None:
            est_locs_df = pd.DataFrame(est_locs)
            est_locs_df.to_excel(writer, sheet_name = app_name+"_locs")

    runtime_df = pd.DataFrame(runtimes, index=[0])
    runtime_df.to_excel(writer, sheet_name = "runtimes")
    writer.save()
    writer.close()
    end = time.time()
    print("Wrote data to:", filepath, np.round(end-start,2), "(sec)")

def MakeSettingPlots(filepath, approaches):
    print(filepath)
    dir_name = os.path.dirname(filepath) + "/"
    file_name = os.path.basename(filepath)
    setting = file_name[:file_name.find("_collection")]
    trials_dict = pd.read_excel(filepath, sheet_name=None, index_col=0)
    full_table = pd.DataFrame()
    for name, sheet in trials_dict.items():
        if "trial" in name:
            full_table = full_table.append(sheet)

    full_table.reset_index(inplace=True, drop=True)

    num_nodes = GetNodeNumFromFilepath(filepath)
    with open('sim_data_params.json') as f:
        sim_data_params = json.load(f)
    spring_model_params = sim_data_params["spring_params"]
    n_init = spring_model_params[4]
    large_data = num_nodes >= sim_data_params["large_data_thresh"]

    max_err_df = None
    avg_err_df = None
    true_pos_df = None
    false_pos_df = None
    runtime_df = None
    for approach in approaches:

        if large_data and "sdp" in approach:
            continue
        if approach == "spring_model":
            app_name = "spring_model_%dinits"%(n_init)
        else:
            app_name = approach

        approach_df = full_table.loc[full_table['estimation_technique'] == app_name]
        max_err_data = approach_df["max_error"].to_numpy()
        max_percent_err_data = approach_df["max_percent_error"].to_numpy()
        avg_err_data = approach_df["avg_error"].to_numpy()
        avg_percent_err_data = approach_df["avg_percent_error"].to_numpy()
        true_pos_data = approach_df["true_pos_rate"].to_numpy()
        false_pos_data = approach_df["false_pos_rate"].to_numpy()
        runtime_data = approach_df["runtime"].to_numpy()

        if runtime_df is None:
            max_err_df = pd.DataFrame(max_err_data, columns=[app_name])
            max_percent_err_df = pd.DataFrame(max_percent_err_data, columns=[app_name])
            avg_err_df = pd.DataFrame(avg_err_data, columns=[app_name])
            avg_percent_err_df = pd.DataFrame(avg_percent_err_data, columns=[app_name])
            true_pos_df = pd.DataFrame(true_pos_data, columns=[app_name])
            false_pos_df = pd.DataFrame(false_pos_data, columns=[app_name])
            runtime_df = pd.DataFrame(runtime_data, columns=[app_name])
        else:
            max_err_df[app_name] = max_err_data
            max_percent_err_df[app_name] = max_percent_err_data
            avg_err_df[app_name]  = avg_err_data
            avg_percent_err_df[app_name]  = avg_percent_err_data
            true_pos_df[app_name]  = true_pos_data
            false_pos_df[app_name]  = false_pos_data
            runtime_df[app_name]  = runtime_data



    book = load_workbook(filepath)
    for sheet in book.sheetnames:
        if  "max_error" in sheet or "max_percent_error" in sheet or "avg_error" in sheet or "avg_percent_error" in sheet or "true_positive" in sheet or "false_positive" in sheet or "runtime" in sheet:
            tag = book.get_sheet_by_name(sheet)
            book.remove_sheet(tag)

    writer = pd.ExcelWriter(filepath, engine = 'openpyxl')
    writer.book = book

    max_err_df.to_excel(writer, sheet_name = "max_error")
    max_percent_err_df.to_excel(writer, sheet_name = "max_percent_error")
    avg_err_df.to_excel(writer, sheet_name = "avg_error")
    avg_percent_err_df.to_excel(writer, sheet_name = "avg_percent_error")
    true_pos_df.to_excel(writer, sheet_name = "true_positive")
    false_pos_df.to_excel(writer, sheet_name = "false_positive")
    runtime_df.to_excel(writer, sheet_name = "runtime")
    writer.save()
    writer.close()

    # Max Error
    # Box-Whiskers
    my_fig, ax = plt.subplots()
    boxplot = max_err_df.boxplot(grid=False)
    figure = plt.gcf() # get current figure
    plt.title("Max. Error")
    plt.ylabel("meters")
    figure.set_size_inches(15, 6)
    my_fig.savefig(dir_name+setting+"_max_err.png", format="png")
    plt.close()

    # Max Error
    # PDF
    max_err_df.plot.kde(ind=1000)
    figure = plt.gcf() # get current figure
    plt.title("PDF of Max Error")
    plt.xlabel("meters")
    plt.xlim(0)
    figure.savefig(dir_name+setting+"_max_err_pdf.png", format="png")
    plt.close()


    # Max Error
    # CDF
    max_err_df = max_err_df.transform(np.sort)
    max_err = max_err_df.max().max()
    offset = 5  # arbitrary buffer
    cols = list(max_err_df)
    xAE = np.linspace(.0, max_err+offset, 100)
    for col in cols:
        errors = max_err_df[col].to_numpy()
        pdeAE_cdf = ndtr(np.subtract.outer(xAE, errors)).mean(axis=1)
        plt.plot(xAE, pdeAE_cdf,label=col)
    plt.xlim(0, max_err+offset)
    plt.title("CDF of Max Error")
    plt.xlabel("meters")
    plt.legend()
    figure = plt.gcf() # get current figure
    figure.savefig(dir_name+setting+"_max_err_cdf.png", format="png")
    plt.close()


    # Max Percent Error
    # Box-Whiskers
    my_fig, ax = plt.subplots()
    boxplot = max_percent_err_df.boxplot(grid=False)
    figure = plt.gcf() # get current figure
    plt.title("Max Percent Error")
    plt.ylabel("meters")
    figure.set_size_inches(15, 6)
    my_fig.savefig(dir_name+setting+"_max_percent_err.png", format="png")
    plt.close()

    # Max Percent Error
    # PDF
    max_percent_err_df.plot.kde(ind=1000)
    figure = plt.gcf() # get current figure
    plt.title("PDF of Max Error")
    plt.xlabel("meters")
    plt.xlim(0)
    figure.savefig(dir_name+setting+"_max_percent_err_pdf.png", format="png")
    plt.close()


    # Max Percent Error
    # CDF
    max_percent_err_df = max_percent_err_df.transform(np.sort)
    max_percent_err = max_percent_err_df.max().max()
    offset = 5  # arbitrary buffer
    cols = list(max_percent_err_df)
    xAE = np.linspace(.0, max_percent_err+offset, 100)
    for col in cols:
        errors = max_percent_err_df[col].to_numpy()
        pdeAE_cdf = ndtr(np.subtract.outer(xAE, errors)).mean(axis=1)
        plt.plot(xAE, pdeAE_cdf,label=col)
    plt.xlim(0, max_percent_err+offset)
    plt.title("CDF of Max Error")
    plt.xlabel("meters")
    plt.legend()
    figure = plt.gcf() # get current figure
    figure.savefig(dir_name+setting+"_max_percent_err_cdf.png", format="png")
    plt.close()


    # Avg Error
    # Box-Whiskers
    my_fig, ax = plt.subplots()
    boxplot = avg_err_df.boxplot(grid=False)
    figure = plt.gcf() # get current figure
    plt.title("Avg. Error")
    plt.ylabel("meters")
    figure.set_size_inches(15, 6)
    my_fig.savefig(dir_name+setting+"_avg_error.png", format="png")
    plt.close()

    # Avg Error
    # PDF
    avg_err_df.plot.kde(ind=1000)
    figure = plt.gcf() # get current figure
    plt.title("PDF of Avg Error")
    plt.xlabel("meters")
    plt.xlim(0)
    figure.savefig(dir_name+setting+"_avg_err_pdf.png", format="png")
    plt.close()


    # Avg Error
    # CDF
    avg_err_df = avg_err_df.transform(np.sort)
    avg_err_max = avg_err_df.max().max()
    offset = 5  # arbitrary buffer
    cols = list(avg_err_df)
    xAE = np.linspace(.0, avg_err_max+offset, 100)
    for col in cols:
        errors = avg_err_df[col].to_numpy()
        pdeAE_cdf = ndtr(np.subtract.outer(xAE, errors)).mean(axis=1)
        plt.plot(xAE, pdeAE_cdf,label=col)
    plt.xlim(0, avg_err_max+offset)
    plt.title("CDF of Avg Error")
    plt.xlabel("meters")
    plt.legend()
    figure = plt.gcf() # get current figure
    figure.savefig(dir_name+setting+"_avg_err_cdf.png", format="png")
    plt.close()


    # Avg Percent Error
    # Box-Whiskers
    my_fig, ax = plt.subplots()
    boxplot = avg_percent_err_df.boxplot(grid=False)
    figure = plt.gcf() # get current figure
    plt.title("Avg. Error")
    plt.ylabel("meters")
    figure.set_size_inches(15, 6)
    my_fig.savefig(dir_name+setting+"_avg_percent_error.png", format="png")
    plt.close()

    # Avg Percent Error
    # PDF
    avg_percent_err_df.plot.kde(ind=1000)
    figure = plt.gcf() # get current figure
    plt.title("PDF of Avg Error")
    plt.xlabel("meters")
    plt.xlim(0)
    figure.savefig(dir_name+setting+"_avg_percent_err_pdf.png", format="png")
    plt.close()


    # Avg Percent Error
    # CDF
    avg_percent_err_df = avg_percent_err_df.transform(np.sort)
    avg_percent_err_max = avg_percent_err_df.max().max()
    offset = 5  # arbitrary buffer
    cols = list(avg_percent_err_df)
    xAE = np.linspace(.0, avg_percent_err_max+offset, 100)
    for col in cols:
        errors = avg_percent_err_df[col].to_numpy()
        pdeAE_cdf = ndtr(np.subtract.outer(xAE, errors)).mean(axis=1)
        plt.plot(xAE, pdeAE_cdf,label=col)
    plt.xlim(0, avg_percent_err_max+offset)
    plt.title("CDF of Avg Error")
    plt.xlabel("meters")
    plt.legend()
    figure = plt.gcf() # get current figure
    figure.savefig(dir_name+setting+"_avg_percent_err_cdf.png", format="png")
    plt.close()


    # True Positive Rate
    my_fig, ax = plt.subplots()
    boxplot = true_pos_df.boxplot(grid=False)
    figure = plt.gcf() # get current figure
    plt.title("True Positive Rate")
    figure.set_size_inches(15, 6)
    my_fig.savefig(dir_name+setting+"_true_positive.png", format="png")
    plt.close()

    # False Positive Rate
    my_fig, ax = plt.subplots()
    boxplot = false_pos_df.boxplot(grid=False)
    figure = plt.gcf() # get current figure
    plt.title("False Positive Rate")
    figure.set_size_inches(15, 6)
    my_fig.savefig(dir_name+setting+"_false_positive.png", format="png")
    plt.close()

    # Runtime
    my_fig, ax = plt.subplots()
    my_fig, ax = plt.subplots()
    boxplot = runtime_df.boxplot(grid=False)
    figure = plt.gcf() # get current figure
    plt.title("Runtime")
    plt.ylabel("seconds")
    figure.set_size_inches(15, 6)
    my_fig.savefig(dir_name+setting+"_runtime.png", format="png")
    plt.close()

    plt.close('all')
    print("Made figures for:", filepath)

if __name__ == '__main__':

    with open('sim_data_params.json') as f:
        sim_data_params = json.load(f)

    snl_approaches = sim_data_params['snl_approaches']
    ble_params = sim_data_params['ble_params']
    pos_contact_thresh = sim_data_params['pos_contact_thresh']
    sim_data_dir = sim_data_params['sim_data_dir']
    trinity_read_dir = sim_data_params['trinity_read_dir']
    trinity_write_dir = sim_data_params['trinity_write_dir']
    collection_dir = sim_data_params['collection_dir']
    num_nodes_list = sim_data_params['num_nodes_list']
    area_lengths = sim_data_params['area_lengths']
    num_repeats = sim_data_params['num_repeat_settings']

    if len(sys.argv) != 3:
        print("Requires two arguments\n")
        print("(1) Options:")
        print("    generate_data")
        print("    perform_snl")
        print("    get_performance_measures")
        print("    gather_performance_data")
        print("    make_plots")
        print()
        print("(2) Options:")
        print("    gauss")
        print("    trinity")
        print()
        print()

    assert (len(sys.argv) == 3)
    mode = sys.argv[1]
    sim_type = sys.argv[2]

    if sim_type == 'gauss':
        data_dir = sim_data_dir
    elif sim_type == 'trinity':
        data_dir = trinity_write_dir
    else:
        raise NotImplementedError

    assert(os.path.isdir(data_dir))
    collection_path = data_dir+collection_dir

    if mode == 'generate_data':
        nproc = multiprocessing.cpu_count()
        pool = multiprocessing.Pool(nproc-2)

        if sim_type == 'gauss':
            for num_nodes in num_nodes_list:
                for area_len in area_lengths:
                    for i in range(num_repeats):
                        # SimulateRssTrial(num_nodes, area_len, i, data_dir, ble_params)
                        pool.apply_async(SimulateRssTrial, args = (num_nodes, area_len, i, data_dir, ble_params))
        elif sim_type == 'trinity':
            for num_nodes in num_nodes_list:
                for area_len in area_lengths:
                    for i in range(num_repeats):
                        # SimulateTrialsFromTrinityData(num_nodes, area_len, i, trinity_read_dir, trinity_write_dir)
                        pool.apply_async(SimulateTrialsFromTrinityData, args = (num_nodes, area_len, i, trinity_read_dir, trinity_write_dir))
        pool.close()
        pool.join()

    # #Perform SNL Techniques
    elif mode == 'perform_snl':
        files = [data_dir+item for item in os.listdir(data_dir) if ".xlsx" in item]
        files.sort(key=lambda file: int(file[file.index("_data/")+6:file.index("nodes_")]))
        nproc = multiprocessing.cpu_count()
        pool = multiprocessing.Pool(nproc-2)
        for f_path in files:
            # TestSNLApproaches(f_path, snl_approaches, ble_params)
            pool.apply_async(TestSNLApproaches, args = (f_path, snl_approaches, ble_params))
        pool.close()
        pool.join()

    # # Get performance measures for SNL techniques
    elif mode == 'get_performance_measures':
        files = [data_dir+item for item in os.listdir(data_dir) if ".xlsx" in item]
        nproc = multiprocessing.cpu_count()
        pool = multiprocessing.Pool(nproc-2)
        for f_path in files:
            # CharacterizePerformance(f_path, snl_approaches, threshold=pos_contact_thresh)
            pool.apply_async(CharacterizePerformance, args = (f_path, snl_approaches, pos_contact_thresh))
        pool.close()
        pool.join()

    # # Get all performance data together and write into collected file
    elif mode == 'gather_performance_data':
        GatherAllTrials(data_dir, collection_path)

    # # Make plots of performance measures
    # # Also group 'trial statistics' into sheets in *collection.xlsx
    elif mode == 'make_plots':
        files = [collection_path+item for item in os.listdir(collection_path) if "collection.xlsx" in item]
        nproc = multiprocessing.cpu_count()
        pool = multiprocessing.Pool(nproc-2)
        for f_path in files:
            MakeSettingPlots(f_path, snl_approaches)
        #     pool.apply_async(MakeSettingPlots, args = (f_path, snl_approaches))
        # pool.close()
        # pool.join()



